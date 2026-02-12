import json
import base64
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

import openpyxl

from .models import SessioneAllestimento, RigaProdotto
from .forms import UploadExcelForm
from core.qr_code_generator import generate_qr_code

# Import per PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm


@login_required
def lista_sessioni(request):
    """Lista delle sessioni di allestimento"""
    sessioni = SessioneAllestimento.objects.all()
    return render(request, 'allestimento/lista_sessioni.html', {
        'sessioni': sessioni
    })


@login_required
def upload_excel(request):
    """Upload di un file Excel per creare una nuova sessione"""
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']

            try:
                # Leggi il file Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                # Estrai info evento dalle prime righe
                nome_evento = ""
                luogo = ""

                # Cerca nelle prime righe per nome evento e luogo
                for row_num in range(1, 8):
                    for col_num in range(1, 6):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value and isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if not nome_evento and cell_value and cell_value != 'Periodo:' and cell_value != 'LISTA MATERIALE':
                                nome_evento = cell_value
                            elif nome_evento and not luogo and cell_value and cell_value != 'Periodo:' and cell_value != 'LISTA MATERIALE':
                                luogo = cell_value
                                break
                    if luogo:
                        break

                # Crea la sessione
                sessione = SessioneAllestimento.objects.create(
                    nome_evento=nome_evento,
                    luogo=luogo,
                    creato_da=request.user
                )

                # Salva il file originale
                file.seek(0)
                sessione.file_originale.save(file.name, file)

                # Estrai le righe prodotto (dalla riga 8 in poi)
                ordine = 0
                for row in ws.iter_rows(min_row=8):
                    quantita = row[0].value
                    descrizione = row[1].value

                    # Salta righe vuote
                    if not quantita and not descrizione:
                        continue

                    if descrizione:
                        ordine += 1
                        RigaProdotto.objects.create(
                            sessione=sessione,
                            ordine=ordine,
                            descrizione=str(descrizione).strip(),
                            quantita_richiesta=int(quantita) if quantita else 0
                        )

                messages.success(request, f'Sessione creata con successo! {ordine} prodotti caricati.')
                return redirect('allestimento:dettaglio_sessione', pk=sessione.pk)

            except Exception as e:
                messages.error(request, f'Errore nel caricamento del file: {str(e)}')
    else:
        form = UploadExcelForm()

    return render(request, 'allestimento/upload_excel.html', {'form': form})


@login_required
def dettaglio_sessione(request, pk):
    """Dettaglio sessione con tabella prodotti"""
    sessione = get_object_or_404(SessioneAllestimento, pk=pk)
    righe = sessione.righe.filter(completata=False)

    # Genera QR code per ogni riga
    righe_con_qr = []
    for riga in righe:
        qr_buffer = generate_qr_code(riga.qr_data, box_size=4, border=2)
        qr_base64 = base64.b64encode(qr_buffer.getvalue()).decode('utf-8')
        righe_con_qr.append({
            'riga': riga,
            'qr_base64': qr_base64
        })

    return render(request, 'allestimento/dettaglio_sessione.html', {
        'sessione': sessione,
        'righe_con_qr': righe_con_qr,
        'righe_completate': sessione.righe.filter(completata=True).count(),
        'righe_totali': sessione.righe.count()
    })


@login_required
@require_POST
def conferma_riga(request, pk):
    """Conferma una singola riga prodotto (AJAX)"""
    riga = get_object_or_404(RigaProdotto, pk=pk)

    try:
        data = json.loads(request.body)
        quantita_allestita = data.get('quantita_allestita', 0)
        note = data.get('note', '')

        riga.quantita_allestita = int(quantita_allestita)
        riga.note = note[:2000]  # Limita a 2000 caratteri
        riga.completata = True
        riga.data_completamento = timezone.now()
        riga.completata_da = request.user
        riga.save()

        # Verifica se tutte le righe sono completate
        sessione = riga.sessione
        tutte_completate = not sessione.righe.filter(completata=False).exists()

        if tutte_completate:
            sessione.completata = True
            sessione.data_completamento = timezone.now()
            sessione.save()

        return JsonResponse({
            'success': True,
            'tutte_completate': tutte_completate,
            'righe_completate': sessione.righe.filter(completata=True).count(),
            'righe_totali': sessione.righe.count()
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_GET
def genera_pdf(request, pk):
    """Genera PDF con tutte le righe completate"""
    sessione = get_object_or_404(SessioneAllestimento, pk=pk)
    righe = sessione.righe.all().order_by('ordine')

    # Buffer per il PDF
    buffer = BytesIO()

    # Documento in landscape per avere più spazio
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Titolo
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#5585b5'),
        spaceAfter=10,
        alignment=1,
    )
    elements.append(Paragraph(f"Report Allestimento - {sessione.nome_evento}", title_style))

    # Sottotitolo con info
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1,
    )
    info_text = f"Luogo: {sessione.luogo} | Data: {sessione.data_creazione.strftime('%d/%m/%Y')}"
    if sessione.data_completamento:
        info_text += f" | Completato: {sessione.data_completamento.strftime('%d/%m/%Y %H:%M')}"
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 20))

    # Stile per celle con testo lungo
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    )

    note_style = ParagraphStyle(
        'NoteStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
    )

    # Header tabella
    headers = ['QR', 'Descrizione', 'Qtà Richiesta', 'Qtà Allestita', 'Note', 'Stato']
    table_data = [headers]

    # Dati tabella
    for riga in righe:
        # Genera QR code come immagine
        qr_buffer = generate_qr_code(riga.qr_data, box_size=2, border=1)
        qr_image = RLImage(qr_buffer, width=1.5*cm, height=1.5*cm)

        # Stato
        if riga.completata:
            stato = "Completato"
            if riga.quantita_allestita != riga.quantita_richiesta:
                stato += f" (diff: {riga.quantita_allestita - riga.quantita_richiesta:+d})"
        else:
            stato = "In attesa"

        # Tronca note per la tabella
        note_display = riga.note[:100] + '...' if len(riga.note) > 100 else riga.note

        row = [
            qr_image,
            Paragraph(riga.descrizione, cell_style),
            str(riga.quantita_richiesta),
            str(riga.quantita_allestita),
            Paragraph(note_display, note_style),
            stato
        ]
        table_data.append(row)

    # Crea tabella con larghezze colonne specifiche
    col_widths = [2*cm, 8*cm, 2.5*cm, 2.5*cm, 6*cm, 3*cm]
    table = Table(table_data, colWidths=col_widths)

    # Stile tabella
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5585b5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),

        # Dati
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # QR centrato
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),  # Quantità centrate
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Stato centrato
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),

        # Bordi
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#5585b5')),

        # Righe alternate
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ])

    # Evidenzia righe con differenze di quantità
    for i, riga in enumerate(righe, start=1):
        if riga.completata and riga.quantita_allestita != riga.quantita_richiesta:
            table_style.add('BACKGROUND', (3, i), (3, i), colors.HexColor('#fff3cd'))
            table_style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#fff3cd'))

    table.setStyle(table_style)
    elements.append(table)

    # Riepilogo finale
    elements.append(Spacer(1, 20))
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        alignment=2,  # Right
    )

    totale_richiesto = sum(r.quantita_richiesta for r in righe)
    totale_allestito = sum(r.quantita_allestita for r in righe)
    differenza = totale_allestito - totale_richiesto

    summary_text = f"<b>Totale richiesto:</b> {totale_richiesto} | "
    summary_text += f"<b>Totale allestito:</b> {totale_allestito} | "
    summary_text += f"<b>Differenza:</b> {differenza:+d}"
    elements.append(Paragraph(summary_text, summary_style))

    # Build PDF
    doc.build(elements)

    # Response
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"allestimento_{sessione.nome_evento.replace(' ', '_')}_{sessione.data_creazione.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def elimina_sessione(request, pk):
    """Elimina una sessione di allestimento"""
    sessione = get_object_or_404(SessioneAllestimento, pk=pk)

    if request.method == 'POST':
        sessione.delete()
        messages.success(request, 'Sessione eliminata con successo.')
        return redirect('allestimento:lista_sessioni')

    return render(request, 'allestimento/conferma_elimina.html', {
        'sessione': sessione
    })
