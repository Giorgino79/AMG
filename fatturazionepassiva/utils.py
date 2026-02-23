# fatturazionepassiva/utils.py

import os
import tempfile
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.utils import timezone
import csv
import json
import logging

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportRiconoscimento:
    """Utility per export riconoscimenti in vari formati"""
    
    def __init__(self, riconoscimento):
        self.riconoscimento = riconoscimento
    
    def export_csv(self):
        """Esporta riconoscimento in CSV"""
        response = HttpResponse(content_type='text/csv')
        filename = f"riconoscimento_{self.riconoscimento.numero_riconoscimento}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Header informazioni generali
        writer.writerow(['RICONOSCIMENTO FORNITORE'])
        writer.writerow(['Numero', self.riconoscimento.numero_riconoscimento])
        writer.writerow(['Fornitore', self.riconoscimento.fornitore.nome])
        writer.writerow(['Periodo', f"{self.riconoscimento.periodo_da} - {self.riconoscimento.periodo_a}"])
        writer.writerow(['Data Creazione', self.riconoscimento.data_creazione.strftime('%d/%m/%Y %H:%M')])
        writer.writerow(['Stato', self.riconoscimento.get_stato_display()])
        writer.writerow([])
        
        # Header righe
        writer.writerow([
            'Prodotto',
            'Codice Interno',
            'EAN', 
            'Origine',
            'Quantità Ordinata',
            'Quantità Riconosciuta',
            'Prezzo Unitario',
            'Aliquota IVA',
            'Totale Imponibile',
            'Totale IVA',
            'Totale con IVA',
            'Descrizione',
            'Note'
        ])
        
        # Righe
        for riga in self.riconoscimento.righe.all():
            writer.writerow([
                riga.prodotto.nome_prodotto,
                riga.prodotto.codice_interno or '',
                riga.prodotto.ean or '',
                riga.get_tipo_origine_display(),
                riga.quantita_ordinata or '',
                riga.quantita_riconosciuta,
                riga.prezzo_unitario,
                f"{riga.aliquota_iva}%",
                riga.get_totale_imponibile(),
                riga.get_totale_iva(),
                riga.get_totale_con_iva(),
                riga.descrizione,
                riga.note
            ])
        
        # Totali
        writer.writerow([])
        writer.writerow(['TOTALI'])
        writer.writerow(['Totale Imponibile', self.riconoscimento.totale_imponibile])
        writer.writerow(['Totale IVA', self.riconoscimento.totale_iva])
        writer.writerow(['Totale Compreso IVA', self.riconoscimento.totale_riconoscimento])
        
        return response
    
    def export_excel(self):
        """Esporta riconoscimento in Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl non disponibile per export Excel")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Riconoscimento"
        
        # Stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Header documento
        ws['A1'] = "RICONOSCIMENTO FORNITORE"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # Informazioni generali
        row = 3
        ws[f'A{row}'] = "Numero:"
        ws[f'B{row}'] = self.riconoscimento.numero_riconoscimento
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Fornitore:"
        ws[f'B{row}'] = self.riconoscimento.fornitore.nome
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Periodo:"
        ws[f'B{row}'] = f"{self.riconoscimento.periodo_da} - {self.riconoscimento.periodo_a}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Stato:"
        ws[f'B{row}'] = self.riconoscimento.get_stato_display()
        ws[f'A{row}'].font = Font(bold=True)
        
        # Spazio
        row += 2
        
        # Header tabella righe
        headers = [
            'Prodotto', 'Codice Interno', 'EAN', 'Origine', 'Qt. Ordinata', 'Qt. Riconosciuta',
            'Prezzo Unit.', 'IVA %', 'Tot. Imponibile', 'Tot. IVA', 'Tot. con IVA', 'Descrizione'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Righe prodotti
        for riga in self.riconoscimento.righe.all():
            row += 1
            ws[f'A{row}'] = riga.prodotto.nome_prodotto
            ws[f'B{row}'] = riga.prodotto.codice_interno or ''
            ws[f'C{row}'] = riga.prodotto.ean or ''
            ws[f'D{row}'] = riga.get_tipo_origine_display()
            ws[f'E{row}'] = float(riga.quantita_ordinata) if riga.quantita_ordinata else None
            ws[f'F{row}'] = float(riga.quantita_riconosciuta)
            ws[f'G{row}'] = float(riga.prezzo_unitario)
            ws[f'H{row}'] = f"{riga.aliquota_iva}%"
            ws[f'I{row}'] = float(riga.get_totale_imponibile())
            ws[f'J{row}'] = float(riga.get_totale_iva())
            ws[f'K{row}'] = float(riga.get_totale_con_iva())
            ws[f'L{row}'] = riga.descrizione
        
        # Totali
        row += 2
        ws[f'I{row}'] = "TOTALE IMPONIBILE:"
        ws[f'J{row}'] = float(self.riconoscimento.totale_imponibile)
        ws[f'I{row}'].font = Font(bold=True)
        ws[f'J{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'I{row}'] = "TOTALE IVA:"
        ws[f'J{row}'] = float(self.riconoscimento.totale_iva)
        ws[f'I{row}'].font = Font(bold=True)
        ws[f'J{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'I{row}'] = "TOTALE GENERALE:"
        ws[f'J{row}'] = float(self.riconoscimento.totale_riconoscimento)
        ws[f'I{row}'].font = Font(bold=True)
        ws[f'J{row}'].font = Font(bold=True)
        
        # Auto-sizing colonne
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva in memoria
        filename = f"riconoscimento_{self.riconoscimento.numero_riconoscimento}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def export_pdf(self):
        """Esporta riconoscimento in PDF"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab non disponibile per export PDF")
        
        filename = f"riconoscimento_{self.riconoscimento.numero_riconoscimento}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Crea PDF
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Titolo
        title = Paragraph(
            f"<b>RICONOSCIMENTO FORNITORE<br/>{self.riconoscimento.numero_riconoscimento}</b>",
            styles['Title']
        )
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Informazioni generali
        info_data = [
            ['Fornitore:', self.riconoscimento.fornitore.nome],
            ['Periodo:', f"{self.riconoscimento.periodo_da} - {self.riconoscimento.periodo_a}"],
            ['Data Creazione:', self.riconoscimento.data_creazione.strftime('%d/%m/%Y %H:%M')],
            ['Stato:', self.riconoscimento.get_stato_display()],
        ]
        
        if self.riconoscimento.note:
            info_data.append(['Note:', self.riconoscimento.note])
        
        info_table = Table(info_data, colWidths=[3*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Tabella righe
        headers = [
            'Prodotto', 'Qt. Riconosc.', 'Prezzo Unit.', 'Tot. Imponibile', 'Tot. IVA', 'Tot. con IVA'
        ]
        
        table_data = [headers]
        
        for riga in self.riconoscimento.righe.all():
            table_data.append([
                riga.prodotto.nome_prodotto[:30],  # Tronca nome lungo
                str(riga.quantita_riconosciuta),
                f"€ {riga.prezzo_unitario}",
                f"€ {riga.get_totale_imponibile()}",
                f"€ {riga.get_totale_iva()}",
                f"€ {riga.get_totale_con_iva()}",
            ])
        
        # Riga totali
        table_data.append([
            'TOTALI', '', '',
            f"€ {self.riconoscimento.totale_imponibile}",
            f"€ {self.riconoscimento.totale_iva}",
            f"€ {self.riconoscimento.totale_riconoscimento}",
        ])
        
        table = Table(table_data, colWidths=[6*cm, 2*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Righe dati
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),
            
            # Riga totali
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (1, -1), (-1, -1), 'RIGHT'),
            
            # Bordi
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Documento generato il {timezone.now().strftime('%d/%m/%Y alle %H:%M')}"
        footer = Paragraph(footer_text, styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        return response


def invia_email_riconoscimento(riconoscimento, email_destinatario, user, includi_allegato=True, tipo_allegato='pdf'):
    """Invia email con riconoscimento al fornitore"""
    
    try:
        # Prepara contesto per template email
        context = {
            'riconoscimento': riconoscimento,
            'fornitore': riconoscimento.fornitore,
            'user': user,
            'data_invio': timezone.now(),
        }
        
        # Genera email da template
        subject = f"Riconoscimento {riconoscimento.numero_riconoscimento} - {riconoscimento.fornitore.nome}"
        
        html_content = render_to_string(
            'fatturazionepassiva/email/riconoscimento_email.html', 
            context
        )
        
        text_content = render_to_string(
            'fatturazionepassiva/email/riconoscimento_email.txt', 
            context
        )
        
        # Crea email
        email = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_destinatario],
        )
        email.attach_alternative(html_content, "text/html")
        
        # Allegato se richiesto
        if includi_allegato:
            exporter = ExportRiconoscimento(riconoscimento)
            
            if tipo_allegato == 'pdf' and REPORTLAB_AVAILABLE:
                response = exporter.export_pdf()
                filename = f"riconoscimento_{riconoscimento.numero_riconoscimento}.pdf"
                email.attach(filename, response.content, 'application/pdf')
                
            elif tipo_allegato == 'excel' and OPENPYXL_AVAILABLE:
                response = exporter.export_excel()
                filename = f"riconoscimento_{riconoscimento.numero_riconoscimento}.xlsx"
                email.attach(filename, response.content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Invia
        email.send()
        
        # Aggiorna riconoscimento
        riconoscimento.inviato_via_email = True
        riconoscimento.data_invio_email = timezone.now()
        riconoscimento.email_destinatario = email_destinatario
        riconoscimento.save()
        
        logger.info(f"Email riconoscimento {riconoscimento.numero_riconoscimento} inviata a {email_destinatario}")
        return True
        
    except Exception as e:
        logger.error(f"Errore invio email riconoscimento {riconoscimento.numero_riconoscimento}: {e}")
        return False